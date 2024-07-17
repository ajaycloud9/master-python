import openpyxl

class ExcelData():
    book = openpyxl.load_workbook("/Users/ajaysingh/Desktop/Python/selenium-pytest/testdata/website-data.xlsx")
    sheet = book.active

    book2 = openpyxl.load_workbook("/Users/ajaysingh/Downloads/download.xlsx")
    sheet2 = book2.active

    @staticmethod
    def exceltodict(testcasename):
        """
        In this example using a static method and
        accessing the class variable using the classname.variable
        """
        data = {}
        for i in range(2,ExcelData.sheet.max_row+1):
            if ExcelData.sheet.cell(row=i, column =1).value == testcasename:
                for j in range(2,ExcelData.sheet.max_column+1):
                    data[ExcelData.sheet.cell(row=1, column =j).value] = ExcelData.sheet.cell(row=i, column =j).value
                return [data]
        return None

    @staticmethod
    def update_fruit_price(fruit,price):
        for i in range(2,ExcelData.sheet2.max_row+1):
            if ExcelData.sheet2.cell(row=i, column =2).value == fruit:
                ExcelData.sheet2.cell(row=i, column =4).value = price
                for j in range(2,ExcelData.sheet2.max_column+1):
                    print(ExcelData.sheet2.cell(row=i, column =j).value)
        ExcelData.book2.save("/Users/ajaysingh/Downloads/download.xlsx")

# ExcelData.update_fruit_price("Mango",2000)