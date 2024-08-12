import openpyxl
book = openpyxl.load_workbook("/Users/ajaysingh/Desktop/Python/selenium-pytest/testdata/website-data.xlsx")


sheet = book.active
cell = sheet.cell(row=1, column =1)
# print(cell.value)
# sheet.cell(row=3, column=3).value = "Rahul"
# print(sheet.cell(row=3, column=3).value)
print(sheet.max_row)
print(sheet.max_column)

# print(sheet['A12'].value)

# for i in range(1,sheet.max_row+1):
#     print(sheet.cell(row=i, column =1).value)
#     for j in range(1,sheet.max_column):
#         print(sheet.cell(row=i, column =j+1).value)
#     print("====")

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.max_column)
        print(sheet.max_row)
        print(sheet.cell(row=i, column =j).value)
    print("====")